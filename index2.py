from openpyxl import Workbook, load_workbook
wb = load_workbook("Excel/003.xlsx")
ws=wb['Sheet1']
def Excel_duoxuanti():
    duoxuanti_mu = []
    duoxuanti_A = []
    duoxuanti_B = []
    duoxuanti_C = []
    duoxuanti_D = []
    duoxuanti_daam = []
    nub = []
    for row in ws.rows:
        nub.append(row)
    for i in range(1, len(nub) + 1):
        duoxuanti_mu.append(ws.cell(row=i, column=1).value)
        duoxuanti_A.append(ws.cell(row=i, column=2).value)
        duoxuanti_B.append(ws.cell(row=i,column=3).value)
        duoxuanti_C.append(ws.cell(row=i,column=4).value)
        duoxuanti_D.append(ws.cell(row=i,column=5).value)
        duoxuanti_daam.append(ws.cell(row=i,column=6).value)
    return duoxuanti_mu,duoxuanti_A,duoxuanti_B,duoxuanti_C,duoxuanti_D,duoxuanti_daam