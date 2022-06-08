from openpyxl import Workbook, load_workbook
wb = load_workbook("Excel/002.xlsx")
ws=wb['Sheet1']
def Excel():
    panduantimu = []
    panduandaan = []
    nub = []
    for row in ws.rows:
        nub.append(row)
    for i in range(1, len(nub) + 1):
        panduantimu.append(ws.cell(row=i, column=1).value)
        panduandaan.append(ws.cell(row=i, column=2).value)
    return panduantimu,panduandaan