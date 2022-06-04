from openpyxl import Workbook, load_workbook
wb = load_workbook("Excel/001.xlsx")
ws=wb['Sheet1']
timu = []
timuA = []
timuB = []
timeC = []
timeD = []
timudan = []
nub = []
# for i in range(0,len(doc.paragraphs)):
#     print(doc.paragraphs[i].text)
def Excel():
    for row in ws.rows:
        nub.append(row)
    for i in range(1,len(nub)+1):
        timu.append(ws.cell(row=i,column =1).value)
        timuA.append(ws.cell(row=i,column = 2).value)
        timuB.append(ws.cell(row=i,column = 3).value)
        timeC.append(ws.cell(row=i,column = 4).value)
        timeD.append(ws.cell(row=i,column = 5).value)
        timudan.append(ws.cell(row=i,column = 6).value)
    return timu,timuA,timuB,timeC,timeD,timudan
